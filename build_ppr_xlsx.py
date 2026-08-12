#!/usr/bin/env python3
"""Convert the Project Progress Report (PPR) PowerPoint into a formatted Excel workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---- shared styles -------------------------------------------------------
def F(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_CL = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# status code -> (label, fill hex, font hex)
STATUS = {
    "C":  ("Complete", "2E75B6", "FFFFFF"),
    "G":  ("On Track", "548235", "FFFFFF"),
    "Y":  ("At Risk", "FFC000", "000000"),
    "R":  ("Significant Risk", "C00000", "FFFFFF"),
    "NS": ("Not Started", "808080", "FFFFFF"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHEAD_FILL = PatternFill("solid", fgColor="D6DCE5")   # light blue-gray
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
LABEL_FILL = PatternFill("solid", fgColor="EDEDED")

NCOLS = 6  # A..F

def set_border(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER

def merge(ws, r, c1, c2, value=None, font=None, fill=None, align=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        for cc in range(c1, c2 + 1):
            ws.cell(row=r, column=cc).fill = fill
    if align:
        cell.alignment = align
    return cell


# =========================================================================
# PPR report sheet builder
# =========================================================================
def build_ppr(wb, title, week_ending, milestones, sponsors, description,
              overall_code, overall_text, metrics, accomplishments,
              next_steps, issues, sheet_name):
    ws = wb.create_sheet(sheet_name)
    widths = [20, 18, 16, 15, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False

    r = 1
    # Title band
    c = merge(ws, r, 1, NCOLS, title, F(20, True, "FFFFFF"), TITLE_FILL, CENTER)
    ws.row_dimensions[r].height = 34
    r += 1
    merge(ws, r, 1, NCOLS, f"Progress week ending:  {week_ending}",
          F(12, True, "FFFFFF"), PatternFill("solid", fgColor="2E4A7D"), CENTER)
    ws.row_dimensions[r].height = 22
    r += 2

    # ---- Milestones table ----
    heads = ["Milestones", "Status", "% Complete", "End Date"]
    merge(ws, r, 1, 3, heads[0], F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    for col, txt in zip([4, 5, 6], heads[1:]):
        ws.cell(row=r, column=col, value=txt).font = F(11, True, "FFFFFF")
        ws.cell(row=r, column=col).fill = HEADER_FILL
        ws.cell(row=r, column=col).alignment = CENTER
    set_border(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 20
    r += 1

    for name, code, pct, end in milestones:
        merge(ws, r, 1, 3, name, F(10), None, WRAP_TL)
        label, fill_hex, font_hex = STATUS[code]
        sc = ws.cell(row=r, column=4, value=code)
        sc.fill = PatternFill("solid", fgColor=fill_hex)
        sc.font = F(10, True, font_hex)
        sc.alignment = CENTER
        pc = ws.cell(row=r, column=5, value=pct)
        pc.font = F(10); pc.alignment = CENTER
        ec = ws.cell(row=r, column=6, value=end)
        ec.font = F(10); ec.alignment = CENTER
        set_border(ws, r, 1, r, NCOLS)
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    # ---- Sponsors ----
    merge(ws, r, 1, NCOLS, "Project Team", F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    set_border(ws, r, 1, r, NCOLS)
    r += 1
    sponsor_rows = [
        ("Executive Sponsor:", sponsors.get("Executive Sponsor", ""),
         "Project Manager:", sponsors.get("Project Manager", "")),
        ("Business Sponsor:", sponsors.get("Business Sponsor", ""), "", ""),
    ]
    for l1, v1, l2, v2 in sponsor_rows:
        ws.cell(row=r, column=1, value=l1).font = F(10, True)
        ws.cell(row=r, column=1).fill = LABEL_FILL
        ws.cell(row=r, column=1).alignment = LEFT
        merge(ws, r, 2, 3, v1, F(10), None, LEFT)
        ws.cell(row=r, column=4, value=l2).font = F(10, True)
        if l2:
            ws.cell(row=r, column=4).fill = LABEL_FILL
        ws.cell(row=r, column=4).alignment = LEFT
        merge(ws, r, 5, 6, v2, F(10), None, LEFT)
        set_border(ws, r, 1, r, NCOLS)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ---- Project Description ----
    merge(ws, r, 1, NCOLS, "Project Description", F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    set_border(ws, r, 1, r, NCOLS)
    r += 1
    merge(ws, r, 1, NCOLS, description, F(10), None, WRAP_TL)
    set_border(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 72
    r += 2

    # ---- Overall Status ----
    merge(ws, r, 1, NCOLS, "Overall Status", F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    set_border(ws, r, 1, r, NCOLS)
    r += 1
    label, fill_hex, font_hex = STATUS[overall_code]
    oc = ws.cell(row=r, column=1, value=overall_code)
    oc.fill = PatternFill("solid", fgColor=fill_hex)
    oc.font = F(11, True, font_hex); oc.alignment = CENTER
    merge(ws, r, 2, NCOLS, overall_text, F(10), None, LEFT)
    set_border(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 20
    r += 2

    # ---- Metrics / KPI table ----
    mheads = ["Description", "Target", "Target Realized to Date", "% of Target Realized"]
    merge(ws, r, 1, 3, mheads[0], F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    for col, txt in zip([4, 5, 6], mheads[1:]):
        ws.cell(row=r, column=col, value=txt).font = F(10, True, "FFFFFF")
        ws.cell(row=r, column=col).fill = HEADER_FILL
        ws.cell(row=r, column=col).alignment = CENTER
    set_border(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 30
    r += 1
    for desc, tgt, realized, pct in metrics:
        merge(ws, r, 1, 3, desc, F(10), None, LEFT)
        for col, val in zip([4, 5, 6], [tgt, realized, pct]):
            cc = ws.cell(row=r, column=col, value=val)
            cc.font = F(10); cc.alignment = CENTER
        set_border(ws, r, 1, r, NCOLS)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ---- Accomplishments / Next Steps / Issues ----
    merge(ws, r, 1, 2, "Accomplishments", F(11, True, "FFFFFF"), HEADER_FILL, CENTER)
    merge(ws, r, 3, 4, "Next Steps", F(11, True, "FFFFFF"), HEADER_FILL, CENTER)
    merge(ws, r, 5, 6, "Issues", F(11, True, "FFFFFF"), HEADER_FILL, CENTER)
    set_border(ws, r, 1, r, NCOLS)
    r += 1
    def bullets(items):
        return "\n".join(f"•  {it}" for it in items) if items else "None"
    merge(ws, r, 1, 2, bullets(accomplishments), F(10), None, WRAP_TL)
    merge(ws, r, 3, 4, bullets(next_steps), F(10), None, WRAP_TL)
    merge(ws, r, 5, 6, bullets(issues), F(10), None, WRAP_TL)
    set_border(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 150
    r += 2

    # ---- Legend ----
    merge(ws, r, 1, NCOLS, "Status Legend", F(10, True, "FFFFFF"), HEADER_FILL, LEFT)
    set_border(ws, r, 1, r, NCOLS)
    r += 1
    legend = [("C", "Complete"), ("G", "On Track"), ("Y", "At Risk"),
              ("R", "Significant Risk w/ inadequate mitigation"), ("NS", "Not Started")]
    col = 1
    start_r = r
    for code, desc in legend:
        _, fill_hex, font_hex = STATUS[code]
        sc = ws.cell(row=r, column=col, value=code)
        sc.fill = PatternFill("solid", fgColor=fill_hex)
        sc.font = F(9, True, font_hex); sc.alignment = CENTER; sc.border = BORDER
        lc = ws.cell(row=r, column=col + 1, value=desc)
        lc.font = F(9); lc.alignment = LEFT; lc.border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 20
    return ws


# =========================================================================
# Build workbook
# =========================================================================
wb = Workbook()
wb.remove(wb.active)

description = ("Augment RCO Patient Financial Navigation resources to increase Medicaid "
              "eligibility identification and coverage capture, thus reducing financial "
              "barriers to care and uncompensated care. Key components include establishing "
              "third party workflows and reporting, expanding coverage identification for "
              "uninsured and underinsured populations, and creating scalable processes that "
              "support both patient financial outcomes and organizational revenue capture.")
sponsors = {"Executive Sponsor": "Michelle Lewis", "Project Manager": "Harrison Nemelka",
            "Business Sponsor": "Rachel Seaman"}

# ---- Report 1: week ending June 26, 2026 ----
milestones1 = [
    ("MSA and SOW fully executed — ElevatePFS partnership formalized (04/10–04/14/2026)", "C", "100%", ""),
    ("Kickoff meeting completed 05/21/2026 — weekly cadence with Elevate team established", "C", "100%", "5/24/2026"),
    ("Transition SP Estimate creation", "G", "90%", "6/29/2026"),
    ("Clearance/ Onboarding", "G", "20%", "7/27/2026"),
    ("Client System Access/Training", "G", "50%", "8/3/2026"),
    ("Create and send reconciliation Demo File", "G", "50%", "7/2/2026"),
    ("Finalization of 90 Day Lookback to send to Elevate", "G", "75%", "6/26/2026"),
    ("Finalize Daily Placement Files to Elevate", "NS", "0%", ""),
    ("Go Live Date", "G", "10%", "8/10/2026"),
]
metrics1 = [
    ("Self Pay to Medicaid Conversion", ">20%", "", ""),
    ("Medicaid Reimbursement", "N/A", "", ""),
    ("Approval Rate", ">75%", "", ""),
    ("Application Turnaround Time", "N/A", "", ""),
]
accomplishments1 = [
    "MSA and SOW fully executed — ElevatePFS partnership formalized (04/10–04/14/2026)",
    "Kickoff meeting completed 05/21/2026 — weekly cadence with Elevate team established",
    "WQ 14856 Insured Inpatient - Prod 6/04",
    "BI 748 PFN Review Complete - Prod 6/04",
    "90 day look back file specs",
    "Epic Access Approved – Template creation beginning",
]
next_steps1 = [
    "Transition Self-Pay Estimate creation - 6/29/2026",
    "Completion of 90-day lookback file",
    "SFTP File Build Testing",
    "Finalize internal team scope",
]
issues1 = ["Elevate Epic Identification - Agency or BI"]

build_ppr(wb, "Medicaid Eligibility Partner", "June 26, 2026", milestones1,
          sponsors, description, "G", "Project is on track to meet milestone timelines.",
          metrics1, accomplishments1, next_steps1, issues1, "PPR - June 26 2026")

# ---- Report 2: week ending July 24, 2026 ----
milestones2 = [
    ("MSA and SOW fully executed — ElevatePFS partnership formalized (04/10–04/14/2026)", "C", "100%", ""),
    ("Kickoff meeting completed 05/21/2026 — weekly cadence with Elevate team established", "C", "100%", "5/24/2026"),
    ("Transition SP Estimate creation", "C", "100%", "6/29/2026"),
    ("Clearance/ Onboarding", "C", "100%", "7/27/2026"),
    ("Client System Access/Training", "C", "100%", "8/3/2026"),
    ("Create and send reconciliation Demo File", "C", "100%", "7/27/2026"),
    ("Finalization of 90 Day Lookback to send to Elevate", "C", "100%", "6/26/2026"),
    ("Finalize Daily Placement Files to Elevate", "C", "100%", "8/10/2026"),
    ("Elevate Notes file", "C", "100%", "7/27/2026"),
    ("Go Live Date", "C", "100%", "8/10/2026"),
]
metrics2 = [
    ("Self Pay to Medicaid Conversion", ">20%", "", ""),
    ("Medicaid Reimbursement", "N/A", "", ""),
    ("Approval Rate", ">75%", "", ""),
    ("Application Turnaround Time", "45 Days", "", ""),
]
accomplishments2 = [
    "MSA and SOW fully executed — ElevatePFS partnership formalized (04/10–04/14/2026)",
    "Kickoff meeting completed 05/21/2026 — weekly cadence with Elevate team established",
    "WQ 14856 Insured Inpatient - Prod 6/04",
    "BI 748 PFN Review Complete - Prod 6/04",
    "90 day look back file specs",
    "Epic access has been approved for Elevate – Template Creation has been finalized",
]
next_steps2 = ["Review and revise plan"]
issues2 = []

build_ppr(wb, "Medicaid Eligibility Partner", "July 24, 2026", milestones2,
          sponsors, description, "G", "Project is on track to meet milestone timelines.",
          metrics2, accomplishments2, next_steps2, issues2, "PPR - July 24 2026")

# =========================================================================
# Sheet: Status Report Key
# =========================================================================
ws = wb.create_sheet("Status Report Key")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110
merge(ws, 1, 1, 2, "Status Report Key", F(16, True, "FFFFFF"), TITLE_FILL, CENTER)
ws.row_dimensions[1].height = 30
ws.cell(row=3, column=1, value="Attribute").font = F(11, True, "FFFFFF")
ws.cell(row=3, column=2, value="Definition").font = F(11, True, "FFFFFF")
for c in (1, 2):
    ws.cell(row=3, column=c).fill = HEADER_FILL
    ws.cell(row=3, column=c).alignment = CENTER
    ws.cell(row=3, column=c).border = BORDER

key_rows = [
    ("Red (R)", "R",
     "One or more of the following conditions exist: Phase/Milestone is at significant risk of delay. "
     "Multiple issues or risks exist, no mitigation plans in place.\n"
     "Schedule: Major date of delivery slippage is expected; > 2-week variance.\n"
     "Resource: Resource availability certain to impact project; >10% variance from projection.\n"
     "Deliverable % Complete: Major deliverables are completed with >2-week variance of planned duration.\n"
     "Budget: Cost variance >5% beyond contingency plan and progress inhibited."),
    ("Yellow (Y)", "Y",
     "One or more of the following conditions exist and none of the above conditions exist: "
     "Phase/Milestone is at risk of missing date of delivery. Active issues or risks exist, "
     "mitigation plan(s) in development.\n"
     "Schedule: ~30% probability minor date of delivery slippage, <2-week variance.\n"
     "Resource: Resource availability may impact date of delivery; 1-9% variance from projection.\n"
     "Deliverable % Complete: Major deliverables and milestones completed on schedule with <2-week variance.\n"
     "Budget: Cost variance >5% beyond contingency plan and progress not yet inhibited or expecting a "
     "cost variance within the next two weeks."),
    ("Green (G)", "G",
     "Project is on track as indicated by all the following conditions existing: "
     "Phase/Milestone is tracking to planned date of delivery. No unmitigated issues or risks.\n"
     "Resource: No resource constraints that will impact date of delivery.\n"
     "Deliverable % Complete: Major deliverables and milestones completed on schedule with <1-week variance.\n"
     "Schedule: Delivery dates are expected to be on time.\n"
     "Budget: No cost variance currently or anticipated within the next two weeks."),
    ("Not Started (NS)", "NS", "Task/deliverable has not started yet."),
    ("Complete (C)", "C", "Task/deliverable is complete."),
]
r = 4
for attr, code, definition in key_rows:
    _, fill_hex, font_hex = STATUS[code]
    ac = ws.cell(row=r, column=1, value=attr)
    ac.fill = PatternFill("solid", fgColor=fill_hex)
    ac.font = F(11, True, font_hex)
    ac.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ac.border = BORDER
    dc = ws.cell(row=r, column=2, value=definition)
    dc.font = F(10); dc.alignment = WRAP_TL; dc.border = BORDER
    ws.row_dimensions[r].height = max(45, 15 * (definition.count("\n") + 1) + 10)
    r += 1

# =========================================================================
# Sheet: Email Instructions
# =========================================================================
ws = wb.create_sheet("Email Instructions")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
merge(ws, 1, 1, 1, "PPR Email Communication Instructions", F(16, True, "FFFFFF"), TITLE_FILL, CENTER)
ws.row_dimensions[1].height = 30

lines = [
    ("Purpose:  Create a consistent and standardized PPR email communication to our stakeholders.", "bold", None),
    ("", None, None),
    ("Steps:", "bold", None),
    ("1.  Copy and paste template to your email.", None, None),
    ("2.  Insert Project Name.", None, None),
    ("3.  Determine appropriate status and delete all others.", None, None),
    ("4.  Add maximum of two impactful accomplishments, next steps, and issues.", None, None),
    ("5.  If you do not have any since the last PPR, list “none”.", None, None),
    ("6.  Insert link to Teams Site.", None, None),
    ("", None, None),
    ("Email Template", "section", None),
    ("Hello,", None, "tmpl"),
    ("", None, "tmpl"),
    ("Please see attached <Insert Project Name Here> Project Report Out (PPR).", None, "tmpl"),
    ("", None, "tmpl"),
    ("Status:  Complete, On Track, At Risk, Significant Risk w/ inadequate mitigation", None, "tmpl"),
    ("", None, "tmpl"),
    ("Accomplishments:", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("", None, "tmpl"),
    ("Next Steps:", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("", None, "tmpl"),
    ("Issues:", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("XX", None, "tmpl"),
    ("", None, "tmpl"),
    ("For additional detailed information, visit the Teams Site. Please reach out with any questions you may have.", None, "tmpl"),
    ("", None, "tmpl"),
    ("Thank you,", None, "tmpl"),
]
TMPL_FILL = PatternFill("solid", fgColor="F2F2F2")
r = 3
for text, style, region in lines:
    cell = ws.cell(row=r, column=1, value=text)
    if style == "bold":
        cell.font = F(11, True)
    elif style == "section":
        merge(ws, r, 1, 1, text, F(11, True, "FFFFFF"), HEADER_FILL, LEFT)
    else:
        cell.font = F(10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if region == "tmpl":
        cell.fill = TMPL_FILL
    r += 1

wb.save("Project_Progress_Report.xlsx")
print("saved Project_Progress_Report.xlsx")
print("sheets:", wb.sheetnames)
